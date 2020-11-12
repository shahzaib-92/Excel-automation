import openpyxl as xl
from openpyxl.chart import PieChart, Reference


def excel_automation(filename):
    # for experiment, ignore it
    # wb = xl.load_workbook("transactions.xlsx")
    wb = xl.load_workbook(f"{filename}")
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)

        # add corrected formula here
        corrected_price = cell.value*0.7

        corrected_price_cell = sheet.cell(row, sheet.max_column + 1)
        corrected_price_cell.value = corrected_price

    graphValues = Reference(
        sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = PieChart()
    chart.add_data(graphValues)

    sheet.add_chart(chart, 'f2')

    wb.save(f'{filename}')
    # for experiment, ignore it
    # wb.save('transactions2.xlsx')
